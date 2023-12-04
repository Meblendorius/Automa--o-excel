import openpyxl
import pyautogui
import keyboard  # Importa a biblioteca keyboard
import mouseinfo

workbook = openpyxl.load_workbook('fonte.xlsx')
vendas_sheet = workbook['Planilha1']
columns = vendas_sheet.max_column
print(columns)

for col in columns:


    for linhas in vendas_sheet.iter_rows(min_row=2):
        pyautogui.click(236, 406, duration=1)
        keyboard.write(str(linhas[0].value))  # Usa keyboard.write para enviar caracteres especiais
        pyautogui.click(296, 406, duration=1)
        keyboard.write(str(linhas[1].value))
        pyautogui.click(362, 406, duration=1)
        keyboard.write(str(linhas[2].value))
        pyautogui.click(424, 406, duration=1)
        keyboard.write(str(linhas[3].value))    